from datetime import datetime as dt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import os
import urllib.parse
import re
import urllib3
urllib3.disable_warnings()
from fake_useragent import UserAgent
ua = UserAgent()
from deep_translator import GoogleTranslator

# Imports for the tracker and the last stopped keyword
import json
from openpyxl import load_workbook
import glob
from selenium.common.exceptions import WebDriverException



keyword_list = ['Transformasi Digital',
                'UKM',
                'Pendapatan',
                'ekspansi bisnis',
                'merger dan akuisisi',
                'pertumbuhan',
                'Investasi eksternal',
                'Investasi',
                'Relokasi',
                'ekspansi global',
                'Digital Transformation',
                'SME',
                'IPO',
                'Revenue',
                'Business Expansion',
                'merger and acquisition',
                'growth',
                'External Investment',
                'invest',
                'Start Up',
                'Relocation',
                'global expansion',
                'listing',
                'audit']

source_list = {'site1' : 'https://www.kontan.co.id', 
               'site2' : 'https://search.bisnis.com', 
               'site3' : 'https://www.cnbcindonesia.com', 
               'site4' : 'https://katadata.co.id', 
               'site5' : 'https://www.detik.com', 
               'site6' : 'https://search.kompas.com',
               'site7' : 'https://www.thejakartapost.com',
               'site8' : 'https://www.liputan6.com' }

remove_term = []

def get_filename():
    now = dt.now()
    filename = "%s-%02d-%02d Crawling Result_ID" % (now.year, now.month, now.day)
    return filename

def site1(keyword): 
    site1_list = []
    print(f"Currently at: {source_list['site1']}, keyword: {keyword}")
    keyword_search = source_list['site1'] + '/search/?search='
    for i in range(1, 3):
        if i==1:
            keyword_search_pg = keyword_search + keyword
        elif i==2:
            keyword_search_pg =  keyword_search + keyword+'&per_page=20'
        elif i==3:
            keyword_search_pg =  keyword_search + keyword+'&per_page=40'
        r = requests.get(keyword_search_pg, verify=False, headers={'User-Agent':ua.chrome})
        soup = bs(r.text, 'lxml')
        article_list = soup.find_all('div', class_='ket')
        for article in article_list:
            title = article.h1.a.text
            if any(oneremoveterm in title for oneremoveterm in remove_term):
                continue
            url = 'https:'+article.h1.a['href']
            try:
                published_on = article.div.text
                to_translate = published_on.split("|")[1]
                published_date = GoogleTranslator(source='auto', target='english').translate(to_translate)
            except:
                published_date = ''
            now = dt.now()
            create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
                
            crawling_one_news = {'Title' : title,
                                'URL' : url,
                                'Publish Date' : published_date,
                                'Keyword' : keyword,
                                'Source' : 'Kontan',
                                'Created Date' : create_date}
            site1_list.append(crawling_one_news)
    return site1_list

def site2(keyword):
    site2_list = []    
    print(f"Currently at: {source_list['site2']}, keyword: {keyword}")
    keyword_search = source_list['site2'] + '/?q='
    for i in range(1, 3):
        keyword_search_pg =  keyword_search + keyword+'&per_page=' + str(i)
        r = requests.get(keyword_search_pg, verify=False, headers={'User-Agent':ua.chrome})
        soup = bs(r.text, 'lxml')
        article_list = soup.find_all('div', class_='col-sm-8')
        for article in article_list:
            title = article.h2.a.text
            if any(oneremoveterm in title for oneremoveterm in remove_term):
                continue
            url = article.h2.a['href']
            try:
                published_on = article.div.div.div.text
                to_translate = published_on.replace("  ","")
                published_date = GoogleTranslator(source='auto', target='english').translate(to_translate)
            except:
                published_date = ''
            now = dt.now()
            create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
            crawling_one_news = {'Title' : title,
                                'URL' : url,
                                'Publish Date' : published_date,
                                'Keyword' : keyword,
                                'Source' : 'Bisnis.com',
                                'Created Date' : create_date}
            site2_list.append(crawling_one_news)
    return site2_list

def site3(keyword):
    site3_list = []
    print(f"Currently at: {source_list['site3']}, keyword: {keyword}")
    keyword_search = source_list['site3'] + '/search?query='
    for i in range(1, 3):
        keyword_search_pg =  keyword_search + keyword+'&p=' + str(i) +'&kanal=&tipe=&date='
        r = requests.get(keyword_search_pg, verify=False, headers={'User-Agent':ua.chrome})
        soup = bs(r.text, 'lxml')
        article_list = soup.find_all('article')
        for article in article_list:
            title = article.a.div.h2.text
            if any(oneremoveterm in title for oneremoveterm in remove_term):
                continue
            url = article.a['href']
            try:
                published_on = article.a.div.span.text
                to_translate = published_on.split("-")[1]
                published_date = GoogleTranslator(source='auto', target='english').translate(to_translate)    
            except:
                published_date = ''    
            now = dt.now()
            create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
            crawling_one_news = {'Title' : title,
                                'URL' : url,
                                'Publish Date' : published_date,
                                'Keyword' : keyword,
                                'Source' : 'CNBC Indonesia',
                                'Created Date' : create_date}
            site3_list.append(crawling_one_news)
    return site3_list

def site4(keyword):
    site4_list = []
    print(f"Currently at: {source_list['site4']}, keyword: {keyword}")
    keyword_search = source_list['site4'] + '/search/news/'
    for i in range(1, 3):
        if i==1:
            search_pg = '0'
        elif i==2:
            search_pg =  '10'
        elif i==3:
            search_pg =  '20'
        keyword_search_pg =  keyword_search + keyword + '/-/-/-/-/-/-/' + search_pg
        r = requests.get(keyword_search_pg, verify=False, headers={'User-Agent':ua.chrome})
        soup = bs(r.text, 'lxml')
        article_list = soup.find_all('article')
        for article in article_list:
            title = article.h3.text
            if any(oneremoveterm in title for oneremoveterm in remove_term):
                continue
            url = article.div.a['href']
            try:
                published_on = article.find('span', class_='article__date').text
                to_translate = published_on.replace(" • ","")
                published_date = GoogleTranslator(source='auto', target='english').translate(to_translate)    
            except:
                published_date = ''    
            now = dt.now()
            create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
            crawling_one_news = {'Title' : title,
                                'URL' : url,
                                'Publish Date' : published_date,
                                'Keyword' : keyword,
                                'Source' : 'katadata',
                                'Created Date' : create_date}
            site4_list.append(crawling_one_news)
    return site4_list

def site5(keyword):
    site5_list = []
    print(f"Currently at: {source_list['site5']}, keyword: {keyword}")
    keyword_search = source_list['site5'] + '/search/searchall?query='
    for i in range(1, 3):
        keyword_search_pg =  keyword_search + keyword +'&siteid=2&sortby=time&page=' + str(i)
        r = requests.get(keyword_search_pg, verify=False, headers={'User-Agent':ua.chrome})
        soup = bs(r.text, 'lxml')
        article_list = soup.find_all('article')
        for article in article_list:
            title = article.find('h2',class_='title').text
            if any(oneremoveterm in title for oneremoveterm in remove_term):
                continue
            url = article.a['href']
            try:
                published_on = article.find('span', class_='date').text
                to_translate = published_on.split(',')[1]
                published_date = GoogleTranslator(source='auto', target='english').translate(to_translate)
            except:
                published_date = ''
            now = dt.now()
            create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
            crawling_one_news = {'Title' : title,
                                'URL' : url,
                                'Publish Date' : published_date,
                                'Keyword' : keyword,
                                'Source' : 'Detik',
                                'Created Date' : create_date}
            site5_list.append(crawling_one_news)
    return site5_list

def site6(keyword):
    site6_list = []
    print(f"Currently at: {source_list['site6']}, keyword: {keyword}")
    keyword_search = source_list['site6'] + '/search/?q='
    for i in range(1, 3):
        keyword_search_pg =  keyword_search + keyword +'&submit=submit#gsc.tab=0&gsc.q=' + keyword +'&gsc.page='+str(i)
        driver = webdriver.Chrome()
        try:
            driver.get(keyword_search_pg)
        except WebDriverException:
            driver.refresh()
        try:
            soup = bs(driver.page_source, 'lxml')
            driver.quit()
            article_list = soup.find_all('div',class_='gs-webResult gs-result')
            for article in article_list:
                title = article.find('div',class_='gs-title').text
                if(title.isspace()):
                    continue
                url = article.a['href']
                try:
                    published_on = article.find('div',class_='gs-bidi-start-align gs-snippet').text
                    published_date = published_on.split('...')[0]    
                except:
                    published_date = ''
                now = dt.now()
                create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
                crawling_one_news = {'Title' : title,
                                    'URL' : url,
                                    'Publish Date' : published_date,
                                    'Keyword' : keyword,
                                    'Source' : 'kompas',
                                    'Created Date' : create_date}
                site6_list.append(crawling_one_news)
        except:
            print("No news found")
            return []

    return site6_list

def site7(keyword):
    site7_list = []
    print(f"Currently at: {source_list['site7']}, keyword: {keyword}")
    keyword_search = source_list['site7'] + '/search?q='
    for i in range(1, 4):
        keyword_search_pg =  keyword_search + keyword +'#gsc.tab=0&gsc.q=' + keyword +'&gsc.page='+str(i)
        driver = webdriver.Chrome()
        try:
            driver.get(keyword_search_pg)
        except WebDriverException:
            driver.refresh()
        soup = bs(driver.page_source, 'lxml')
        driver.quit()
        article_list = soup.find_all('div',class_='gs-webResult gs-result')
        for article in article_list:
            title = article.find('div',class_='gs-title').text
            if(title.isspace()):
                continue
            url = article.a['href']
            try:
                published_on = article.find('div',class_='gs-bidi-start-align gs-snippet').text
                published_date = published_on.split('...')[0]    
            except:
                published_date = ''
            now = dt.now()
            create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
            crawling_one_news = {'Title' : title,
                                'URL' : url,
                                'Publish Date' : published_date,
                                'Keyword' : keyword,
                                'Source' : 'The Jakarta Post',
                                'Created Date' : create_date}
            site7_list.append(crawling_one_news)
    return site7_list

def site8(keyword):
    site8_list = []
    print(f"Currently at: {source_list['site8']}, keyword: {keyword}")
    keyword_search_pg = source_list['site8']
    driver = webdriver.Chrome()
    try:
        driver.get(keyword_search_pg)
    except WebDriverException:
        driver.refresh()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class,'top')]//input[@name='q']"))).send_keys(keyword)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,"//div[contains(@class,'top')]//button[@type='submit']"))).click()
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@class='search-results__header']"))).click()
    soup = bs(driver.page_source, 'lxml')
    driver.quit()
    try:
        article_list = soup.find_all('article')
        i = 0
        for article in article_list:
            if(i > 30):
                break
            i = i+1    
            try:
                title = article.h4.a['title']
            except:
                continue
            url = article.h4.a['href']
            try:
                published_date = article.time['title']
            except:
                published_date = ''
                continue
            now = dt.now()
            create_date = "%s-%02d-%02d %02d:%02d:%02d" % (now.year, now.month, now.day, now.hour, now.minute, now.second)
            crawling_one_news = {'Title' : title,
                                'URL' : url,
                                'Publish Date' : published_date,
                                'Keyword' : keyword,
                                'Source' : 'Liputan6',
                                'Created Date' : create_date}
            site8_list.append(crawling_one_news)
    except:
        print("No news found")
        return []
    return site8_list

def get_today_news():
    filepath = os.path.dirname(os.path.abspath('__file__'))
    filename = os.path.join(filepath, 'excel', get_filename() + '.xlsx')
    print("Gathering Keyword News.......")
    
    # Last stopped keyword VARIABLE
    lastStoppedKeyword = ""
    # Stored excel name VARIABLE
    storedExcelName = ""

    # Generate the txt file if it does not exist for the very first time
    if not os.path.exists("Indonesia_Keywords_Tracker.txt"):
        with open("Indonesia_Keywords_Tracker.txt", 'w') as f:
            pass  # Do nothing here, just create the file   

    # if file not empty, store last stopped keyword
    if os.path.getsize('Indonesia_Keywords_Tracker.txt') != 0:
        with open("Indonesia_Keywords_Tracker.txt", "r") as f:
            my_dict_str = f.read()
            my_dict = json.loads(my_dict_str)
            print(my_dict)
            lastStoppedKeyword = my_dict["stoppedAtKeyword"]
            storedExcelName = my_dict["storedExcelName"]
            print("Last stopped keyword: " + lastStoppedKeyword)
            print("Stored Excel Name: " + storedExcelName)
        
        # once stored, wipe the txt file
        with open("Indonesia_Keywords_Tracker.txt", "w") as f:
            pass

    # ================ Check if lastStoppedKeyword is empty ================
    if lastStoppedKeyword == "":
        print("lastStoppedKeyword is empty! Starting from the top of the keyword list") 
        for keyword in keyword_list:
            # Restart newslist in fresh state
            newslist = []

            try:
                # Get the keyword for each site store it inside newslist
                newslist += site1(keyword) + site2(keyword) + site3(keyword) + site4(keyword) + site5(keyword) + site6(keyword) + site7(keyword) + site8(keyword)

                # Create the excel file of the date ran if it does not exist, the problem of the filename being on the next day does not really occur here because filename is initiated at the top of the function (So even if 11.55pm we scrape, filename will still be may-17 instead of may-18)

                # If there is an error, store the keyword and the excel name in the txt file, in the next iteration the elif will be executed, and we will start from the last stopped keyword and that particular excel file
                
                if not os.path.isfile(filename):
                    df = pd.DataFrame(newslist)
                    df = df.drop_duplicates('Title')
                    df.to_excel(filename, index=False)
                    print("Created Excel File with keyword: " + keyword)

                else:
                # Else, get the latest modified excel file from excel local directory, because if we webscrape from 11:55pm to 12am, date changes. Because of that we want to get the latest modified excel file
                     # Sort the files by modified date, and get the latest file
                    latest_file = max(glob.iglob(os.path.join(filepath, 'excel', '*.xlsx')), key=os.path.getmtime)
                    df_existing= pd.read_excel(latest_file, engine="openpyxl", sheet_name='Sheet1')
                    # append the newslist into the excel file
                    df_new = pd.DataFrame(newslist)
                    # Combine the existing and new newslist dataframes
                    combined_data = df_existing.append(df_new, ignore_index=True)
                    # Drop the duplicates
                    combined_data = combined_data.drop_duplicates('Title')
                    print("duplicates dropped")
                    print("Appending... using OpenPyxl")
                    # Append to latest excel file
                    with pd.ExcelWriter(latest_file, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
                        # Using openpyxl to append the dataframe into the existing excel file
                        #writer.book = load_workbook(latest_file)
                        combined_data.to_excel(writer, sheet_name='Sheet1', index=False, header=True)
                    print("Appended to Excel File with keyword: " + keyword)
            except:
                #  ==== Scenario 1: Crawling fails IF lastStoppedKeyword is empty ====
               # Store the failed keyword here; This means that we need to start over from this keyword again.
                    keywordToAdd = keyword
                    print("Error at " + keyword + " keyword. Stopping here....")
                    # Store the failed keyword into a txt file, in a dictionary format, with keyword and excel name stored in the txt file
                    with open("Indonesia_Keywords_Tracker.txt", "w") as f:
                        data = {"stoppedAtKeyword": keywordToAdd, "storedExcelName": filename}
                        f.write(json.dumps(data))
                        f.close()
                        return ("Restart the program again to continue from where you left off. This is stored in the tracker txt. Keyword: " + keywordToAdd, "Excel Name: " + filename)
                    

   
    # ================ If the file is not empty and if its the same month ================
    elif lastStoppedKeyword != "":
        print("lastStoppedKeyword is not empty! Starting from the last stopped keyword: " + lastStoppedKeyword)
            
        # Get the last stopped keyword
        lastStoppedKeyword = my_dict["stoppedAtKeyword"]
        # Get the stored excel name
        storedExcelName = my_dict["storedExcelName"]
        # Get the index of the last stopped keyword
        lastStoppedKeywordIndex = keyword_list.index(lastStoppedKeyword)
        # Get the remaining keywords from the last stopped keyword
        remainingKeywordList = keyword_list[lastStoppedKeywordIndex:]

        # Wipe the txt file
        with open("Indonesia_Keywords_Tracker.txt", "w") as f:
            pass

         # For each remaining keyword, get the news and store it inside newslist
        for keyword in remainingKeywordList:
            # Restart newslist in fresh state
            newslist = []
            try:
                # Get the keyword for each site store it inside newslist
                newslist += site1(keyword) + site2(keyword) + site3(keyword) + site4(keyword) + site5(keyword) + site6(keyword) + site7(keyword) + site8(keyword)
                # Assume that it already exists, so we just need to append here, since we already have the tracker and the keyword
                # Read from the stored excel file
                df_existing= pd.read_excel(storedExcelName, engine="openpyxl", sheet_name='Sheet1')
                # append the newslist into the excel file
                df_new = pd.DataFrame(newslist)
                # Combine the existing and new newslist dataframes
                combined_data = df_existing.append(df_new, ignore_index=True)
                # Drop the duplicates
                combined_data = combined_data.drop_duplicates('Title')
                print("duplicates dropped")
                print("Appending... using OpenPyxl")
                # Append to latest excel file

                with pd.ExcelWriter(storedExcelName, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
                    # Using openpyxl to append the dataframe into the existing excel file
                    #writer.book = load_workbook(storedExcelName)
                    combined_data.to_excel(writer, sheet_name='Sheet1', index=False, header=True)
                print("Appended to Excel File with keyword: " + keyword)

            except:
                # ==== Scenario 2: Crawling fails IF lastStoppedKeyword is not empty ====
                # Store the failed keyword here; This means that we need to start over from this keyword again.
                keywordToAdd = keyword
                print("Error at " + keyword + " keyword. Stopping here....")
                # Store the failed keyword into a txt file, in a dictionary format, with keyword and date
                with open("Indonesia_Keywords_Tracker.txt", "w") as f:
                    data = {"stoppedAtKeyword": keywordToAdd, "storedExcelName": storedExcelName}
                    f.write(json.dumps(data))
                    f.close()
                    return ("Restart the program again to continue from where you left off. This is stored in the tracker txt. Keyword: " + lastStoppedKeyword, "Excel Name: " + storedExcelName)
                

    try:
        notInExcel = ""
        df = pd.read_excel(filename)
        # keywords that are not in excel file, print the remaining out
        for keyword in keyword_list:
            # Read the excel file
            # If keyword is not in the excel file, print it out

            if keyword not in df['Keyword'].values:
                notInExcel += keyword + " "

        # Print out the keywords that are not in the excel file
        print("Successfully iterated through all keywords!" )
        # Total length of articles in the excel file
        print("Total length of articles in the excel file: " + str(len(df)))
    except:
        print("No excel has been created yet. Please run the program again to create the excel file")

    
  

def main():
    print("Start Exporting!")
    get_today_news()


if __name__ == "__main__":
    main()
    